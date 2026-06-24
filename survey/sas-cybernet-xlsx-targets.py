#!/usr/bin/env python3
"""Read-only Cybernet xlsx target ingester for SysAdminSuite (offline, openpyxl)."""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("[sas-cybernet-xlsx] ERROR: openpyxl required (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)

MANIFEST_FIELDS = ["Identifier", "IdentifierType", "DeviceType", "HostName", "Serial", "MACAddress", "Source"]
REPORT_FIELDS = [
    "InputSerial", "InputHostName", "ResolvedHostName", "ResolvedSerial",
    "ResolvedMACAddress", "ResolutionStatus", "Source", "GapNote",
]
GAP_FIELDS = [
    "Identifier", "IdentifierType", "HostName", "Serial", "MACAddress",
    "ResolutionStatus", "GapReason", "Source",
]
Rec = dict[str, str]
EMPTY = {"", "N/A", "NA", "NONE", "NULL", "-", "--", "TBD", "UNKNOWN", "#N/A", "#REF!"}
HOSTNAME_RE = re.compile(r"^[A-Za-z]{2,6}\d{2,}[A-Za-z0-9_-]*$|^[A-Za-z0-9]+[-_][A-Za-z0-9]+")
MAC_RE = re.compile(r"(?i)(?:[0-9a-f]{2}[:-]){5}[0-9a-f]{2}|[0-9a-f]{12}")
HOST_HEADERS = ("cybernet hostname", "pc name", "hostname", "host name", "host", "computer name")
SERIAL_HEADERS = ("cybernet serial", "pc / cybernet serial no", "cybernet serial number", "serial number", "serial", "service tag")
MAC_HEADERS = ("cybernet mac", "mac address", "mac")
NEURON_MAC_HEADERS = ("neuron mac",)


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() in EMPTY else text


def norm_serial(value: object) -> str:
    text = re.sub(r"\s+", "", clean(value)).upper()
    return text if len(text) >= 3 else ""


def norm_host(value: object) -> str:
    text = re.sub(r"\s+", "", clean(value)).upper()
    return text if text and HOSTNAME_RE.search(text) else ""


def norm_mac(value: object) -> str:
    text = clean(value)
    if not text:
        return ""
    match = MAC_RE.search(text)
    if not match:
        return ""
    raw = re.sub(r"[^0-9A-Fa-f]", "", match.group(0)).upper()
    return ":".join(raw[i : i + 2] for i in range(0, 12, 2)) if len(raw) == 12 else ""


def identifier_type(value: str) -> str:
    value = clean(value)
    if not value:
        return "Unknown"
    hexv = re.sub(r"[^0-9A-Fa-f]", "", value)
    if len(hexv) == 12 and re.search(r"[:\-.]|^[0-9A-Fa-f]{12}$", value):
        return "MAC"
    return "HostName" if HOSTNAME_RE.search(value) else "Serial"


def merge_source(*parts: str) -> str:
    seen: list[str] = []
    for part in parts:
        for item in str(part or "").split(";"):
            item = item.strip()
            if item and item not in seen:
                seen.append(item)
    return ";".join(seen)


def merge_pair(left: Rec, right: Rec) -> Rec:
    out = left.copy()
    for field in ("host", "serial", "mac"):
        if not out.get(field) and right.get(field):
            out[field] = right[field]
    out["source"] = merge_source(out.get("source", ""), right.get("source", ""))
    return out


def records_overlap(left: Rec, right: Rec) -> bool:
    return any(left.get(k) and left[k] == right.get(k) for k in ("serial", "host"))


def merge_records(records: list[Rec]) -> list[Rec]:
    merged = [rec.copy() for rec in records if any(rec.get(k) for k in ("host", "serial", "mac"))]
    changed = True
    while changed:
        changed, groups, used = False, [], [False] * len(merged)
        for index, rec in enumerate(merged):
            if used[index]:
                continue
            current, used[index] = rec.copy(), True
            for other_index in range(index + 1, len(merged)):
                if used[other_index] or not records_overlap(current, merged[other_index]):
                    continue
                current = merge_pair(current, merged[other_index])
                used[other_index] = changed = True
            groups.append(current)
        merged = groups
    return merged


def resolution_status(rec: Rec) -> str:
    count = sum(1 for key in ("host", "serial", "mac") if rec.get(key))
    return "FULL" if count >= 3 else "PARTIAL" if count >= 2 else "MINIMAL"


def gap_reason(rec: Rec) -> str:
    if resolution_status(rec) == "FULL":
        return ""
    missing = [name for key, name in (("host", "HostName"), ("serial", "Serial"), ("mac", "MACAddress")) if not rec.get(key)]
    return f"missing:{','.join(missing)}" if missing else "minimal_identity"


def cell(row: tuple[object, ...], index: int | None) -> str:
    return clean(row[index]) if index is not None and index < len(row) else ""


def header_map(row: tuple[object, ...]) -> dict[str, int]:
    return {
        re.sub(r"\s+", " ", str(value or "").strip()).lower(): idx
        for idx, value in enumerate(row)
        if clean(value)
    }


def pick_column(headers: dict[str, int], names: tuple[str, ...]) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    for key, idx in headers.items():
        if any(name in key for name in names):
            return idx
    return None


def record_from_row(row: tuple[object, ...], headers: dict[str, int], source: str, *, neuron_mac: bool = False) -> Rec | None:
    host_idx = pick_column(headers, HOST_HEADERS)
    serial_idx = pick_column(headers, SERIAL_HEADERS)
    mac_idx = pick_column(headers, NEURON_MAC_HEADERS if neuron_mac else MAC_HEADERS)
    if mac_idx is None and neuron_mac:
        mac_idx = pick_column(headers, MAC_HEADERS)
    host, serial, mac = norm_host(cell(row, host_idx)), norm_serial(cell(row, serial_idx)), norm_mac(cell(row, mac_idx))
    return None if not (host or serial or mac) else {"host": host, "serial": serial, "mac": mac, "source": source}


def parse_alejandro_workbook(path: Path) -> list[Rec]:
    records: list[Rec] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            title, upper = ws.title.strip(), ws.title.strip().upper()
            source = f"Alejandro:{path.name}:{title}"
            if "AKBAR WAVE" in upper:
                records.extend({"host": "", "serial": s, "mac": "", "source": source}
                               for row in ws.iter_rows(values_only=True)
                               if (s := norm_serial(cell(row, 0))))
            elif upper.startswith("PO"):
                for row in ws.iter_rows(values_only=True):
                    host, serial = norm_host(cell(row, 0)), norm_serial(cell(row, 1))
                    if host or serial:
                        records.append({"host": host, "serial": serial, "mac": "", "source": source})
    finally:
        wb.close()
    return records


def parse_enrichment_sheet(ws, path: Path, title: str) -> list[Rec]:
    lower, source, records = title.lower(), f"Enrichment:{path.name}:{title}", []
    if lower == "ssuh host":
        return [{"host": h, "serial": "", "mac": "", "source": source}
                for row in ws.iter_rows(values_only=True) if (h := norm_host(cell(row, 0)))]
    headers, serial_idx = None, 1
    for row in ws.iter_rows(values_only=True):
        row_headers = header_map(row)
        if lower == "cdw stock":
            if "cybernet serial number" in row_headers:
                headers, serial_idx = row_headers, row_headers["cybernet serial number"]
                continue
            if headers and (serial := norm_serial(cell(row, serial_idx))):
                records.append({"host": "", "serial": serial, "mac": "", "source": source})
            continue
        if lower == "neuron cybernet":
            if "pc name" in row_headers and "pc / cybernet serial no" in row_headers:
                headers = row_headers
                continue
            if headers and (rec := record_from_row(row, headers, source, neuron_mac=True)):
                records.append(rec)
            continue
        if any(name in row_headers for name in HOST_HEADERS + SERIAL_HEADERS + MAC_HEADERS):
            headers = row_headers
            continue
        if headers and (rec := record_from_row(row, headers, source)):
            records.append(rec)
    return records


def parse_enrichment_workbook(path: Path) -> list[Rec]:
    records: list[Rec] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            title, lower = ws.title.strip(), ws.title.strip().lower()
            if lower in {"ssuh host", "cdw stock", "neuron cybernet", "deployments", "ssuh configs"} or lower.startswith("configured cybernets"):
                records.extend(parse_enrichment_sheet(ws, path, title))
    finally:
        wb.close()
    return records


def enrich_records(base: list[Rec], enrichment: list[Rec]) -> list[Rec]:
    by_serial, by_host = {}, {}
    for rec in enrichment:
        if rec.get("serial"):
            by_serial.setdefault(rec["serial"], []).append(rec)
        if rec.get("host"):
            by_host.setdefault(rec["host"], []).append(rec)
    enriched = []
    for rec in base:
        out = rec.copy()
        matches = []
        if out.get("serial"):
            matches.extend(by_serial.get(out["serial"], []))
        if out.get("host"):
            matches.extend(by_host.get(out["host"], []))
        for match in matches:
            for field in ("host", "serial", "mac"):
                if not out.get(field) and match.get(field):
                    out[field] = match[field]
            out["source"] = merge_source(out.get("source", ""), match.get("source", ""))
        enriched.append(out)
    return enriched


def match_before_state(after: Rec, before: list[Rec]) -> Rec:
    matched, hits = {"host": "", "serial": "", "mac": "", "source": ""}, 0
    for candidate in before:
        if records_overlap(after, candidate):
            matched, hits = merge_pair(matched, candidate), hits + 1
    return matched if hits else after.copy()


def manifest_row(rec: Rec, device_type: str) -> Rec:
    host, serial, mac = rec.get("host", ""), rec.get("serial", ""), rec.get("mac", "")
    identifier = host or serial or mac
    itype = identifier_type(identifier)
    if not host and itype == "HostName":
        host = norm_host(identifier)
    if not serial and itype == "Serial":
        serial = norm_serial(identifier)
    if not mac and itype == "MAC":
        mac = norm_mac(identifier)
    return {
        "Identifier": identifier, "IdentifierType": itype, "DeviceType": device_type,
        "HostName": host, "Serial": serial, "MACAddress": mac, "Source": rec.get("source", ""),
    }


def write_csv(path: Path, fields: list[str], rows: list[Rec]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, quoting=csv.QUOTE_ALL, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def build_reports(before: list[Rec], after: list[Rec], device_type: str) -> tuple[list[Rec], list[Rec]]:
    report, gaps = [], []
    for src, dst in zip(before, after):
        status, note = resolution_status(dst), gap_reason(dst)
        report.append({
            "InputSerial": src.get("serial", ""), "InputHostName": src.get("host", ""),
            "ResolvedHostName": dst.get("host", ""), "ResolvedSerial": dst.get("serial", ""),
            "ResolvedMACAddress": dst.get("mac", ""), "ResolutionStatus": status,
            "Source": dst.get("source", ""), "GapNote": note,
        })
        if status != "FULL":
            manifest = manifest_row(dst, device_type)
            gaps.append({
                "Identifier": manifest["Identifier"], "IdentifierType": manifest["IdentifierType"],
                "HostName": manifest["HostName"], "Serial": manifest["Serial"],
                "MACAddress": manifest["MACAddress"], "ResolutionStatus": status,
                "GapReason": note or "minimal_identity", "Source": manifest["Source"],
            })
    return report, gaps


def main() -> int:
    parser = argparse.ArgumentParser(description="Read-only Cybernet xlsx target ingester")
    parser.add_argument("--workbook", required=True, help="Primary Alejandro-style workbook")
    parser.add_argument("--enrichment", action="append", default=[], help="Enrichment workbook (repeatable)")
    parser.add_argument("--output", default="survey/output/cybernet_alejandro_targets.csv")
    parser.add_argument("--report", default="survey/output/cybernet_alejandro_enrichment_report.csv")
    parser.add_argument("--gaps", default="survey/output/cybernet_alejandro_gaps.csv")
    parser.add_argument("--device-type", default="Cybernet")
    args = parser.parse_args()

    workbook = Path(args.workbook)
    if not workbook.is_file():
        print(f"[sas-cybernet-xlsx] ERROR: workbook not found: {workbook}", file=sys.stderr)
        return 1

    base = parse_alejandro_workbook(workbook)
    if not base:
        print(f"[sas-cybernet-xlsx] ERROR: no targets parsed from {workbook}", file=sys.stderr)
        return 1

    enrichment_rows: list[Rec] = []
    for path_str in args.enrichment:
        path = Path(path_str)
        if not path.is_file():
            print(f"[sas-cybernet-xlsx] ERROR: enrichment workbook not found: {path}", file=sys.stderr)
            return 1
        enrichment_rows.extend(parse_enrichment_workbook(path))

    merged_before = merge_records(base)
    enriched = enrich_records(merged_before, enrichment_rows) if enrichment_rows else [r.copy() for r in merged_before]
    merged_after = merge_records(enriched)
    manifest = [manifest_row(rec, args.device_type) for rec in merged_after]
    report, gaps = build_reports([match_before_state(rec, merged_before) for rec in merged_after], merged_after, args.device_type)

    output, report_path, gaps_path = Path(args.output), Path(args.report), Path(args.gaps)
    write_csv(output, MANIFEST_FIELDS, manifest)
    write_csv(report_path, REPORT_FIELDS, report)
    write_csv(gaps_path, GAP_FIELDS, gaps)

    full = sum(row["ResolutionStatus"] == "FULL" for row in report)
    partial = sum(row["ResolutionStatus"] == "PARTIAL" for row in report)
    minimal = sum(row["ResolutionStatus"] == "MINIMAL" for row in report)
    print(f"[sas-cybernet-xlsx] manifest={len(manifest)} report={len(report)} gaps={len(gaps)} FULL={full} PARTIAL={partial} MINIMAL={minimal}")
    print(f"[sas-cybernet-xlsx] wrote {output}")
    print(f"[sas-cybernet-xlsx] wrote {report_path}")
    print(f"[sas-cybernet-xlsx] wrote {gaps_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
