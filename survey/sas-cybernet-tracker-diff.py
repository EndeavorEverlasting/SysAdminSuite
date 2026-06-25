#!/usr/bin/env python3
"""Compare Alejandro Cybernet serials against the deployment tracker inventory."""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("[sas-cybernet-tracker-diff] ERROR: openpyxl required (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)

EMPTY = {"", "N/A", "NA", "NONE", "NULL", "-", "--", "TBD", "UNKNOWN", "#N/A", "#REF!"}
MAC_RE = re.compile(r"(?i)(?:[0-9a-f]{2}[:-]){5}[0-9a-f]{2}|[0-9a-f]{12}")
HOSTNAME_RE = re.compile(r"^[A-Za-z]{2,6}\d{2,}[A-Za-z0-9_-]*$|^[A-Za-z0-9]+[-_][A-Za-z0-9]+")
HEADER_TOKENS = {
    "serial", "serial number", "cybernet serial", "cybernet serial number",
    "pc / cybernet serial no", "service tag",
    "host", "hostname", "host name", "cybernet host", "cybernet hostname",
    "computer name", "pc name",
    "mac", "mac address", "cybernet mac", "cybernet mac address",
    "neuron mac", "neuron mac address", "neuron s/n", "neuron serial", "neuron serial number",
    "device type", "deployed",
}
MANIFEST_FIELDS = ["Identifier", "IdentifierType", "DeviceType", "HostName", "Serial", "MACAddress", "Source"]
ALEJANDRO_FIELDS = ["Serial", "RowCount", "HostNames", "Sources", "ProbeReady"]
TRACKER_FIELDS = ["Serial", "TrackerRowCount", "DeployedYesCount", "HostNames", "MACAddresses", "Sources"]
TRACKED_FIELDS = [
    "Serial", "AlejandroRowCount", "AlejandroHostNames", "TrackerRowCount",
    "TrackerDeployedYesCount", "TrackerHostNames", "TrackerMACAddresses", "Sources",
]
DUP_FIELDS = [
    "IdentifierKind", "Identifier", "DeployedYesCount", "TrackerRowCount",
    "HostNames", "Serials", "MACAddresses", "NeuronMACAddresses", "NeuronSerials", "Sources",
]
DUP_IDENTIFIER_KINDS = ("host", "serial", "mac", "neuron_mac", "neuron_sn")
# Serial-first progress buckets. Denominator is always unique Alejandro serials,
# never hostname rows. Optional evidence CSVs are enrichment only and never
# upgrade a serial to confirmed on their own (ping/AD are not serial proof).
PROGRESS_FIELDS = [
    "TotalSerialTargets",
    "SurveyedSerials",
    "RemainingSerials",
    "HostResolvedSerials",
    "SerialOnlyReviewRequired",
    "AmbiguousHostnameSerials",
    "ADCandidateSerials",
    "PingReachableCandidates",
    "NeedsPrivilegedIdentity",
    "PercentComplete",
]
PROGRESS_META_FIELDS = ["PopulationAuthority", "GeneratedAt"]
REACHABLE_TOKENS = {"REACHABLE", "UP", "ONLINE", "YES", "OPEN"}


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() in EMPTY else text


def norm_serial(value: object) -> str:
    text = re.sub(r"\s+", "", clean(value)).upper()
    return text if len(text) >= 3 else ""


def norm_host(value: object) -> str:
    text = re.sub(r"\s+", "", clean(value)).upper()
    return text if text and HOSTNAME_RE.search(text) else ""


def norm_mac(value: object) -> str:
    match = MAC_RE.search(clean(value))
    if not match:
        return ""
    raw = re.sub(r"[^0-9A-Fa-f]", "", match.group(0)).upper()
    return ":".join(raw[i : i + 2] for i in range(0, 12, 2)) if len(raw) == 12 else ""


def norm_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def is_header_label(value: object) -> bool:
    return norm_header(value) in HEADER_TOKENS


def joined(values: set[str]) -> str:
    return ";".join(sorted(v for v in values if v))


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, quoting=csv.QUOTE_ALL, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def parse_alejandro(path: Path) -> dict[str, dict[str, object]]:
    serials: dict[str, dict[str, object]] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            title = ws.title.strip()
            upper = title.upper()
            if "AKBAR WAVE" in upper:
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    cell = row[0] if row else ""
                    if is_header_label(cell):
                        continue
                    serial = norm_serial(cell)
                    if serial:
                        add_alejandro(serials, serial, "", f"{path.name}:{title}:R{row_index}")
            elif upper.startswith("PO"):
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    host_cell = row[0] if len(row) > 0 else ""
                    serial_cell = row[1] if len(row) > 1 else ""
                    if is_header_label(host_cell) or is_header_label(serial_cell):
                        continue
                    host = norm_host(host_cell)
                    serial = norm_serial(serial_cell)
                    if serial:
                        add_alejandro(serials, serial, host, f"{path.name}:{title}:R{row_index}")
    finally:
        wb.close()
    return serials


def add_alejandro(serials: dict[str, dict[str, object]], serial: str, host: str, source: str) -> None:
    rec = serials.setdefault(serial, {"count": 0, "hosts": set(), "sources": set()})
    rec["count"] = int(rec["count"]) + 1
    rec["hosts"].add(host)
    rec["sources"].add(source)


def find_header_row(rows: list[tuple[object, ...]], scan_limit: int) -> tuple[int, dict[str, int]]:
    aliases = {
        "deployed": "deployed",
        "device type": "device_type",
        "cybernet hostname": "host",
        "cybernet host": "host",
        "hostname": "host",
        "computer name": "host",
        "cybernet serial": "serial",
        "cybernet serial number": "serial",
        "serial number": "serial",
        "serial": "serial",
        "cybernet mac": "mac",
        "cybernet mac address": "mac",
        "mac address": "mac",
        "mac": "mac",
        "neuron mac": "neuron_mac",
        "neuron mac address": "neuron_mac",
        "neuron s/n": "neuron_sn",
        "neuron serial": "neuron_sn",
        "neuron serial number": "neuron_sn",
    }
    for index, row in enumerate(rows[:scan_limit]):
        mapped: dict[str, int] = {}
        for col_index, value in enumerate(row):
            key = aliases.get(norm_header(value))
            if key and key not in mapped:
                mapped[key] = col_index
        if "serial" in mapped and ("host" in mapped or "deployed" in mapped):
            return index, mapped
    return -1, {}


def row_value(row: tuple[object, ...], index: int | None) -> str:
    return clean(row[index]) if index is not None and index < len(row) else ""


def parse_tracker(path: Path, sheet_name: str, header_scan_rows: int) -> tuple[dict[str, dict[str, object]], list[dict[str, str]]]:
    serials: dict[str, dict[str, object]] = {}
    identifier_rows: list[dict[str, str]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    header_index, columns = find_header_row(rows, header_scan_rows)
    if header_index < 0:
        raise ValueError(f"could not find tracker headers in first {header_scan_rows} rows")

    for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        host = norm_host(row_value(row, columns.get("host")))
        serial = norm_serial(row_value(row, columns.get("serial")))
        mac = norm_mac(row_value(row, columns.get("mac")))
        neuron_mac = norm_mac(row_value(row, columns.get("neuron_mac")))
        neuron_sn = norm_serial(row_value(row, columns.get("neuron_sn")))
        deployed = row_value(row, columns.get("deployed")).upper() == "YES"
        if not (host or serial or mac or neuron_mac or neuron_sn):
            continue
        source = f"{path.name}:{sheet_name}:R{offset}"
        if serial:
            add_tracker_serial(serials, serial, host, mac, deployed, source)
        identifier_rows.append({
            "host": host, "serial": serial, "mac": mac,
            "neuron_mac": neuron_mac, "neuron_sn": neuron_sn,
            "deployed": "YES" if deployed else "NO", "source": source,
        })
    return serials, identifier_rows


def add_tracker_serial(serials: dict[str, dict[str, object]], serial: str, host: str, mac: str, deployed: bool, source: str) -> None:
    rec = serials.setdefault(serial, {"count": 0, "deployed": 0, "hosts": set(), "macs": set(), "sources": set()})
    rec["count"] = int(rec["count"]) + 1
    rec["deployed"] = int(rec["deployed"]) + (1 if deployed else 0)
    rec["hosts"].add(host)
    rec["macs"].add(mac)
    rec["sources"].add(source)


def duplicate_exceptions(identifier_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in identifier_rows:
        for kind in DUP_IDENTIFIER_KINDS:
            value = row.get(kind, "")
            if value:
                grouped[(kind, value)].append(row)

    exceptions = []
    for (kind, value), rows in sorted(grouped.items()):
        deployed_count = sum(1 for row in rows if row["deployed"] == "YES")
        if deployed_count <= 1:
            continue
        exceptions.append({
            "IdentifierKind": kind,
            "Identifier": value,
            "DeployedYesCount": str(deployed_count),
            "TrackerRowCount": str(len(rows)),
            "HostNames": joined({row["host"] for row in rows}),
            "Serials": joined({row["serial"] for row in rows}),
            "MACAddresses": joined({row["mac"] for row in rows}),
            "NeuronMACAddresses": joined({row.get("neuron_mac", "") for row in rows}),
            "NeuronSerials": joined({row.get("neuron_sn", "") for row in rows}),
            "Sources": joined({row["source"] for row in rows}),
        })
    return exceptions


def alejandro_rows(alejandro: dict[str, dict[str, object]]) -> list[dict[str, str]]:
    rows = []
    for serial, rec in sorted(alejandro.items()):
        hosts = sorted(h for h in rec["hosts"] if h)
        rows.append({
            "Serial": serial,
            "RowCount": str(rec["count"]),
            "HostNames": joined(set(hosts)),
            "Sources": joined(rec["sources"]),
            # probe-ready only when exactly one resolved hostname; multiple
            # hostnames stay review-required rather than crowning one.
            "ProbeReady": "Yes" if len(hosts) == 1 else "No",
        })
    return rows


def tracker_rows(tracker: dict[str, dict[str, object]]) -> list[dict[str, str]]:
    return [{
        "Serial": serial,
        "TrackerRowCount": str(rec["count"]),
        "DeployedYesCount": str(rec["deployed"]),
        "HostNames": joined(rec["hosts"]),
        "MACAddresses": joined(rec["macs"]),
        "Sources": joined(rec["sources"]),
    } for serial, rec in sorted(tracker.items())]


def already_tracked_rows(alejandro: dict[str, dict[str, object]], tracker: dict[str, dict[str, object]]) -> list[dict[str, str]]:
    rows = []
    for serial in sorted(set(alejandro) & set(tracker)):
        arec, trec = alejandro[serial], tracker[serial]
        rows.append({
            "Serial": serial,
            "AlejandroRowCount": str(arec["count"]),
            "AlejandroHostNames": joined(arec["hosts"]),
            "TrackerRowCount": str(trec["count"]),
            "TrackerDeployedYesCount": str(trec["deployed"]),
            "TrackerHostNames": joined(trec["hosts"]),
            "TrackerMACAddresses": joined(trec["macs"]),
            "Sources": joined(arec["sources"] | trec["sources"]),
        })
    return rows


def untracked_manifest(alejandro: dict[str, dict[str, object]], tracker: dict[str, dict[str, object]], device_type: str) -> list[dict[str, str]]:
    rows = []
    for serial in sorted(set(alejandro) - set(tracker)):
        rec = alejandro[serial]
        hosts = sorted(h for h in rec["hosts"] if h)
        source = joined(rec["sources"])
        if len(hosts) == 1:
            # exactly one resolved hostname: probe-ready
            hostname = hosts[0]
            identifier, identifier_type = hostname, "HostName"
        else:
            # zero or multiple hostnames: do not crown one; keep serial-keyed and
            # not probe-ready. Surface ambiguous candidates for human review.
            hostname = ""
            identifier, identifier_type = serial, "Serial"
            if len(hosts) > 1:
                source = f"{source};review:ambiguous_hostnames={'|'.join(hosts)}"
        rows.append({
            "Identifier": identifier,
            "IdentifierType": identifier_type,
            "DeviceType": device_type,
            "HostName": hostname,
            "Serial": serial,
            "MACAddress": "",
            "Source": source,
        })
    return rows


def read_csv_optional(path_value: str | None) -> list[dict[str, str]]:
    if not path_value:
        return []
    path = Path(path_value)
    if not path.is_file():
        print(f"[sas-cybernet-tracker-diff] WARN: optional evidence not found: {path}", file=sys.stderr)
        return []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def cell(row: dict[str, str], *names: str) -> str:
    lowered = {str(key).lower(): value for key, value in row.items() if key}
    for name in names:
        value = lowered.get(name.lower())
        if value not in (None, ""):
            return clean(value)
    return ""


def is_reachable(value: str) -> bool:
    return value.strip().upper() in REACHABLE_TOKENS


def is_identity_collected(value: str) -> bool:
    return "IDENTITYCOLLECTED" in re.sub(r"\s+", "", value).upper()


def render_progress_bar(percent: float, width: int = 20) -> str:
    filled = max(0, min(width, round(width * percent / 100)))
    return f"[{'#' * filled}{'-' * (width - filled)}] {percent:.1f}%"


def build_progress_summary(
    alejandro: dict[str, dict[str, object]],
    tracker: dict[str, dict[str, object]],
    identity_rows: list[dict[str, str]],
    preflight_rows: list[dict[str, str]],
    ad_rows: list[dict[str, str]],
) -> dict[str, object]:
    total = len(alejandro)
    tracked = set(alejandro) & set(tracker)

    observed_serials: set[str] = set()
    reachable_hosts: set[str] = set()
    identity_collected_hosts: set[str] = set()
    for row in identity_rows:
        host = norm_host(cell(row, "ObservedHostName", "HostName", "DnsName"))
        target_host = norm_host(cell(row, "Target", "HostName"))
        ping = cell(row, "PingStatus")
        status = cell(row, "IdentityStatus")
        observed_serial = norm_serial(cell(row, "ObservedSerial", "Serial"))
        for key in {host, target_host}:
            if not key:
                continue
            if is_reachable(ping):
                reachable_hosts.add(key)
            if is_identity_collected(status):
                identity_collected_hosts.add(key)
        if observed_serial:
            observed_serials.add(observed_serial)

    for row in preflight_rows:
        host = norm_host(cell(row, "Target", "HostName"))
        if host and is_reachable(cell(row, "PingStatus")):
            reachable_hosts.add(host)

    ad_serials: set[str] = set()
    ad_hosts: set[str] = set()
    for row in ad_rows:
        ad_serial = norm_serial(cell(row, "Serial", "ObservedSerial", "ExpectedSerial"))
        ad_host = norm_host(cell(row, "HostName", "Host", "Name", "DNSHostName"))
        if ad_serial:
            ad_serials.add(ad_serial)
        if ad_host:
            ad_hosts.add(ad_host)

    host_resolved = serial_only = ambiguous = 0
    ad_candidates = ping_candidates = needs_identity = 0
    surveyed: set[str] = set(tracked)

    for serial, rec in alejandro.items():
        hosts = sorted(h for h in rec["hosts"] if h)
        if len(hosts) == 1:
            host_resolved += 1
        elif not hosts:
            serial_only += 1
        else:
            ambiguous += 1

        # Identity is the only optional signal that can mark an untracked serial
        # surveyed; ping/AD raise candidate confidence but never confirm.
        confirmed = serial in observed_serials or any(h in identity_collected_hosts for h in hosts)
        if serial not in tracker and confirmed:
            surveyed.add(serial)

        if serial in ad_serials or any(h in ad_hosts for h in hosts):
            ad_candidates += 1

        if any(h in reachable_hosts for h in hosts):
            ping_candidates += 1
            if serial not in surveyed:
                needs_identity += 1

    surveyed_count = len(surveyed)
    remaining = total - surveyed_count
    percent = round(100 * surveyed_count / total, 1) if total else 0.0

    summary: dict[str, object] = {
        "TotalSerialTargets": total,
        "SurveyedSerials": surveyed_count,
        "RemainingSerials": remaining,
        "HostResolvedSerials": host_resolved,
        "SerialOnlyReviewRequired": serial_only,
        "AmbiguousHostnameSerials": ambiguous,
        "ADCandidateSerials": ad_candidates,
        "PingReachableCandidates": ping_candidates,
        "NeedsPrivilegedIdentity": needs_identity,
        "PercentComplete": percent,
        "PopulationAuthority": "alejandro_serials",
        "GeneratedAt": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    return summary


def format_progress_line(summary: dict[str, object]) -> str:
    bar = render_progress_bar(float(summary["PercentComplete"]))
    return (
        f"{bar} "
        f"{summary['SurveyedSerials']}/{summary['TotalSerialTargets']} serials surveyed | "
        f"{summary['RemainingSerials']} remaining | "
        f"{summary['NeedsPrivilegedIdentity']} need identity | "
        f"{summary['AmbiguousHostnameSerials']} ambiguous"
    )


def write_progress_json(path: Path, summary: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, indent=2)
        handle.write("\n")


def progress_csv_row(summary: dict[str, object]) -> dict[str, str]:
    return {field: str(summary.get(field, "")) for field in PROGRESS_FIELDS + PROGRESS_META_FIELDS}


def main() -> int:
    parser = argparse.ArgumentParser(description="Diff Alejandro Cybernet serials against deployment tracker serial inventory")
    parser.add_argument("--alejandro", required=True, help="Alejandro-style Cybernet workbook")
    parser.add_argument("--tracker", required=True, help="Deployment tracker workbook")
    parser.add_argument("--tracker-sheet", default="Deployments")
    parser.add_argument("--output-prefix", default="survey/output/cybernet")
    parser.add_argument("--device-type", default="Cybernet")
    parser.add_argument("--header-scan-rows", type=int, default=40)
    parser.add_argument("--identity-csv", help="optional workstation_identity.csv (enrichment evidence; only IdentityCollected confirms a serial)")
    parser.add_argument("--preflight-csv", help="optional network_preflight.csv (ping reachability; not serial proof)")
    parser.add_argument("--ad-serial-csv", help="optional AD live-serial export (enrichment candidates; not serial proof)")
    parser.add_argument("--no-progress", action="store_true", help="suppress the tech-visible progress bar line")
    args = parser.parse_args()

    alejandro_path, tracker_path = Path(args.alejandro), Path(args.tracker)
    if not alejandro_path.is_file():
        print(f"[sas-cybernet-tracker-diff] ERROR: Alejandro workbook not found: {alejandro_path}", file=sys.stderr)
        return 1
    if not tracker_path.is_file():
        print(f"[sas-cybernet-tracker-diff] ERROR: tracker workbook not found: {tracker_path}", file=sys.stderr)
        return 1

    try:
        alejandro = parse_alejandro(alejandro_path)
        tracker, identifier_rows = parse_tracker(tracker_path, args.tracker_sheet, args.header_scan_rows)
    except ValueError as exc:
        print(f"[sas-cybernet-tracker-diff] ERROR: {exc}", file=sys.stderr)
        return 1

    prefix = Path(args.output_prefix)
    outputs = {
        f"{prefix}_alejandro_unique_serials.csv": (ALEJANDRO_FIELDS, alejandro_rows(alejandro)),
        f"{prefix}_tracker_unique_serials.csv": (TRACKER_FIELDS, tracker_rows(tracker)),
        f"{prefix}_alejandro_already_tracked.csv": (TRACKED_FIELDS, already_tracked_rows(alejandro, tracker)),
        f"{prefix}_alejandro_untracked.csv": (MANIFEST_FIELDS, untracked_manifest(alejandro, tracker, args.device_type)),
        f"{prefix}_tracker_duplicate_exceptions.csv": (DUP_FIELDS, duplicate_exceptions(identifier_rows)),
    }
    for path_str, (fields, rows) in outputs.items():
        write_csv(Path(path_str), fields, rows)
        print(f"[sas-cybernet-tracker-diff] wrote {path_str} rows={len(rows)}")

    summary = build_progress_summary(
        alejandro,
        tracker,
        read_csv_optional(args.identity_csv),
        read_csv_optional(args.preflight_csv),
        read_csv_optional(args.ad_serial_csv),
    )
    progress_json = f"{prefix}_progress_summary.json"
    progress_csv = f"{prefix}_progress_summary.csv"
    write_progress_json(Path(progress_json), summary)
    write_csv(Path(progress_csv), PROGRESS_FIELDS + PROGRESS_META_FIELDS, [progress_csv_row(summary)])
    print(f"[sas-cybernet-tracker-diff] wrote {progress_json}")
    print(f"[sas-cybernet-tracker-diff] wrote {progress_csv}")

    print(
        "[sas-cybernet-tracker-diff] "
        f"alejandro_unique_serials={len(alejandro)} "
        f"tracker_unique_serials={len(tracker)} "
        f"already_tracked={len(set(alejandro) & set(tracker))} "
        f"untracked={len(set(alejandro) - set(tracker))}"
    )
    if not args.no_progress:
        print(f"[sas-cybernet-tracker-diff] {format_progress_line(summary)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
