#!/usr/bin/env python3
"""Sanitized fixture builder and serial-comparison contract helpers for Bash contract tests."""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("openpyxl required", file=sys.stderr)
    sys.exit(1)

EMPTY = {"", "N/A", "NA", "NONE", "NULL", "-", "--", "TBD", "UNKNOWN", "#N/A", "#REF!"}
SERIAL_HEADERS = (
    "cybernet serial",
    "pc / cybernet serial no",
    "cybernet serial number",
    "serial number",
    "serial",
    "service tag",
)
HOST_HEADERS = ("cybernet hostname", "pc name", "hostname", "host name", "host", "computer name")
MAC_HEADERS = ("cybernet mac", "mac address", "mac")
IDENTIFIER_FIELDS = ("Cybernet Serial", "Cybernet Hostname", "Cybernet MAC", "Neuron MAC", "Neuron S/N")
HEADER_TOKENS = {
    "serial", "serial number", "cybernet serial", "cybernet serial number",
    "pc / cybernet serial no", "service tag",
    "host", "hostname", "host name", "cybernet host", "cybernet hostname",
    "computer name", "pc name",
    "mac", "mac address", "cybernet mac", "cybernet mac address",
    "neuron mac", "neuron mac address", "neuron s/n", "neuron serial", "neuron serial number",
    "device type", "deployed",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() in EMPTY else text


def norm_serial(value: object) -> str:
    text = re.sub(r"\s+", "", clean(value)).upper()
    return text if len(text) >= 3 else ""


def norm_host(value: object) -> str:
    text = re.sub(r"\s+", "", clean(value)).upper()
    hostname_re = re.compile(r"^[A-Za-z]{2,6}\d{2,}[A-Za-z0-9_-]*$|^[A-Za-z0-9]+[-_][A-Za-z0-9]+")
    return text if text and hostname_re.search(text) else ""


def norm_mac(value: object) -> str:
    text = clean(value)
    if not text:
        return ""
    match = re.search(r"(?i)(?:[0-9a-f]{2}[:-]){5}[0-9a-f]{2}|[0-9a-f]{12}", text)
    if not match:
        return ""
    raw = re.sub(r"[^0-9A-Fa-f]", "", match.group(0)).upper()
    return ":".join(raw[i : i + 2] for i in range(0, 12, 2)) if len(raw) == 12 else ""


def norm_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def is_header_label(value: object) -> bool:
    return norm_header(value) in HEADER_TOKENS


def norm_identifier(field: str, value: object) -> str:
    field_lower = field.strip().lower()
    if "serial" in field_lower or field_lower.endswith("s/n"):
        return norm_serial(value)
    if "mac" in field_lower:
        return norm_mac(value)
    if "hostname" in field_lower or field_lower in {"pc name", "host name", "host", "computer name"}:
        return norm_host(value)
    return clean(value).upper()


def is_deployed_yes(value: object) -> bool:
    return clean(value).upper() == "YES"


def header_map(row: tuple[object, ...]) -> dict[str, int]:
    return {
        re.sub(r"\s+", " ", str(value or "").strip()).lower(): idx
        for idx, value in enumerate(row)
        if clean(value)
    }


def pick_column(headers: dict[str, int], names: tuple[str, ...]) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    for key, idx in headers.items():
        if any(name in key for name in names):
            return idx
    return None


def parse_alejandro_serials(path: Path) -> list[str]:
    serials: list[str] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            upper = ws.title.strip().upper()
            if "AKBAR WAVE" in upper:
                for row in ws.iter_rows(values_only=True):
                    cell = row[0] if row else ""
                    if is_header_label(cell):
                        continue
                    if serial := norm_serial(cell):
                        serials.append(serial)
            elif upper.startswith("PO"):
                for row in ws.iter_rows(values_only=True):
                    host_cell = row[0] if len(row) > 0 else ""
                    serial_cell = row[1] if len(row) > 1 else ""
                    if is_header_label(host_cell) or is_header_label(serial_cell):
                        continue
                    if serial := norm_serial(serial_cell):
                        serials.append(serial)
    finally:
        wb.close()
    return serials


def parse_tracker_rows(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            if ws.title.strip().lower() != "deployments":
                continue
            headers = None
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row_headers = header_map(row)
                if any(name in row_headers for name in HOST_HEADERS + SERIAL_HEADERS + MAC_HEADERS):
                    headers = row_headers
                    continue
                if not headers:
                    continue
                deployed_idx = headers.get("deployed")
                record = {
                    "excel_row": str(row_index),
                    "sheet": ws.title,
                    "deployed_yes": "yes" if is_deployed_yes(row[deployed_idx] if deployed_idx is not None and deployed_idx < len(row) else "") else "no",
                }
                for label in IDENTIFIER_FIELDS:
                    key = label.lower()
                    idx = headers.get(key)
                    if idx is None:
                        for header_name, header_idx in headers.items():
                            if key in header_name:
                                idx = header_idx
                                break
                    value = row[idx] if idx is not None and idx < len(row) else ""
                    record[label] = norm_identifier(label, value)
                if any(record.get(label) for label in IDENTIFIER_FIELDS):
                    rows.append(record)
    finally:
        wb.close()
    return rows


def unique_serial_inventory(serials: list[str]) -> list[str]:
    seen: list[str] = []
    for serial in serials:
        if serial and serial not in seen:
            seen.append(serial)
    return seen


def compare_serial_inventories(alejandro_serials: list[str], tracker_rows: list[dict[str, str]]) -> dict[str, list[str]]:
    alejandro_unique = unique_serial_inventory(alejandro_serials)
    tracker_serials = [row.get("Cybernet Serial", "") for row in tracker_rows if row.get("Cybernet Serial")]
    tracker_unique = unique_serial_inventory(tracker_serials)
    alejandro_set = set(alejandro_unique)
    tracker_set = set(tracker_unique)
    return {
        "alejandro_unique_serials": alejandro_unique,
        "tracker_unique_serials": tracker_unique,
        "already_tracked": sorted(alejandro_set & tracker_set),
        "untracked": sorted(alejandro_set - tracker_set),
    }


def duplicate_exceptions(tracker_rows: list[dict[str, str]]) -> tuple[list[dict[str, object]], dict[str, int]]:
    non_deployed_repeats: dict[str, int] = {}
    grouped: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in tracker_rows:
        for field in IDENTIFIER_FIELDS:
            value = row.get(field, "")
            if not value:
                continue
            grouped.setdefault((field, value), []).append(row)
            if row.get("deployed_yes") != "yes":
                non_deployed_repeats[value] = non_deployed_repeats.get(value, 0) + 1

    exceptions: list[dict[str, object]] = []
    for (field, value), hits in sorted(grouped.items()):
        deployed_hits = [row for row in hits if row.get("deployed_yes") == "yes"]
        if len(deployed_hits) <= 1:
            continue
        exceptions.append(
            {
                "identifier_field": field,
                "identifier": value,
                "deployed_yes_count": len(deployed_hits),
                "matching_row_count": len(hits),
                "rows": ";".join(sorted({row["excel_row"] for row in deployed_hits}, key=int)),
                "sheet": hits[0].get("sheet", ""),
            }
        )
    return exceptions, non_deployed_repeats


def build_duplicate_fixture(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AKBAR WAVE 1"
    ws["A1"] = "MEDTEST24-DUP01"
    ws["A2"] = "MEDTEST24-DUP01"
    ws["A3"] = "MEDTEST24-UNIQ01"
    po = wb.create_sheet("PO 1")
    po["A1"] = "wts001opr001"
    po["B1"] = "MEDTEST24-DUP01"
    wb.save(path)


def build_diff_fixture(primary: Path, tracker: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AKBAR WAVE 1"
    ws["A1"] = "MEDTEST24-TRACKED01"
    ws["A2"] = "MEDTEST24-NEW01"
    wb.save(primary)

    ewb = Workbook()
    dep = ewb.active
    dep.title = "Deployments"
    dep.append(["Device Type", "Cybernet Hostname", "Cybernet Serial", "Cybernet MAC", "Neuron S/N", "Deployed"])
    dep.append(["Cybernet-Neuron", "WTS001OPR101", "MEDTEST24-TRACKED01", "000D050AA101", "NEU-UNIQ01", "Yes"])
    dep.append(["Cybernet-Neuron", "WTS001OPR201", "MEDTEST24-DUPYES01", "000D050AA201", "NEU-DUP01", "Yes"])
    dep.append(["Cybernet-Neuron", "WTS001OPR202", "MEDTEST24-DUPYES01", "000D050AA202", "NEU-DUP01", "Yes"])
    dep.append(["Cybernet-Neuron", "WTS001OPR103", "MEDTEST24-HIST01", "000D050AA103", "NEU-HIST01", "No"])
    dep.append(["Cybernet-Neuron", "WTS001OPR104", "MEDTEST24-HIST01", "000D050AA104", "NEU-HIST01", "No"])
    ewb.save(tracker)


def build_serial_only_fixture(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AKBAR WAVE 9"
    ws["A1"] = "MEDTEST24-SERIALONLY01"
    wb.save(path)


def build_header_row_fixture(path: Path) -> None:
    """Alejandro workbook whose first rows are header labels, not serials."""
    wb = Workbook()
    ws = wb.active
    ws.title = "AKBAR WAVE 1"
    ws["A1"] = "Cybernet Serial"
    ws["A2"] = "MEDTEST24-HDR01"
    po = wb.create_sheet("PO 1")
    po["A1"] = "Cybernet Hostname"
    po["B1"] = "Cybernet Serial"
    po["A2"] = "WTS001OPR301"
    po["B2"] = "MEDTEST24-HDR02"
    wb.save(path)


def build_ambiguous_host_fixture(path: Path) -> None:
    """One Alejandro serial mapped to two distinct hostnames (review-required)."""
    wb = Workbook()
    po = wb.active
    po.title = "PO 1"
    po["A1"] = "WTS001OPR401"
    po["B1"] = "MEDTEST24-AMBIG01"
    po["A2"] = "WTS001OPR402"
    po["B2"] = "MEDTEST24-AMBIG01"
    wb.save(path)


def emit_fixtures(output_dir: Path) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "primary_dup": str(output_dir / "alejandro-dup-serial.xlsx"),
        "primary_diff": str(output_dir / "alejandro-diff.xlsx"),
        "tracker_diff": str(output_dir / "tracker-diff.xlsx"),
        "primary_serial_only": str(output_dir / "alejandro-serial-only.xlsx"),
        "primary_header_rows": str(output_dir / "alejandro-header-rows.xlsx"),
        "primary_ambiguous": str(output_dir / "alejandro-ambiguous-host.xlsx"),
    }
    build_duplicate_fixture(Path(paths["primary_dup"]))
    build_diff_fixture(Path(paths["primary_diff"]), Path(paths["tracker_diff"]))
    build_serial_only_fixture(Path(paths["primary_serial_only"]))
    build_header_row_fixture(Path(paths["primary_header_rows"]))
    build_ambiguous_host_fixture(Path(paths["primary_ambiguous"]))
    return paths


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--emit-fixtures", metavar="DIR")
    parser.add_argument("--validate-diff", nargs=2, metavar=("ALEJANDRO", "TRACKER"))
    parser.add_argument("--validate-duplicates", metavar="TRACKER")
    args = parser.parse_args()

    if args.emit_fixtures:
        print(json.dumps(emit_fixtures(Path(args.emit_fixtures))))
        return 0

    if args.validate_diff:
        alejandro_path, tracker_path = map(Path, args.validate_diff)
        result = compare_serial_inventories(
            parse_alejandro_serials(alejandro_path),
            parse_tracker_rows(tracker_path),
        )
        print(json.dumps(result))
        return 0

    if args.validate_duplicates:
        tracker_rows = parse_tracker_rows(Path(args.validate_duplicates))
        exceptions, non_deployed = duplicate_exceptions(tracker_rows)
        print(json.dumps({"duplicate_exceptions": exceptions, "non_deployed_repeats": non_deployed}))
        return 0

    parser.error("no action requested")
    return 2


if __name__ == "__main__":
    sys.exit(main())
